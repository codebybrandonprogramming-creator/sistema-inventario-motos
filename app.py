
# Conexión a MySQL para manejo real en producción

from flask import Flask, request, render_template, redirect, url_for, flash, send_file, session, jsonify
import json
import os
from datetime import datetime
from collections import defaultdict
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import pymysql
from flask_wtf import CSRFProtect
import openpyxl
import pandas as pd


app = Flask(__name__)

# Filtro personalizado para formato colombiano
@app.template_filter('formato_colombiano')
def formato_colombiano(valor, decimales=3):
    """
    Formatea números al estilo colombiano: punto para miles, coma para decimales
    Ejemplo: 95000.436 -> "95.000,436"
    """
    if valor is None:
        valor = 0
    
    # Formatear con los decimales especificados
    formato = f"{{:,.{decimales}f}}"
    texto = formato.format(float(valor))
    
    # Cambiar coma por punto (miles) y punto por coma (decimales)
    texto = texto.replace(',', 'TEMP').replace('.', ',').replace('TEMP', '.')
    
    return texto

app.secret_key = "dev_secret_key_change_in_production"

# Headers de seguridad
@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    return response

# ---------------------------------------------------------------------------------
# CONFIGURACIÓN Y CONEXIÓN A MYSQL
# ---------------------------------------------------------------------------------
def get_db_connection():
    return pymysql.connect(
        host=os.getenv('DB_HOST', '127.0.0.1'),
        user=os.getenv('DB_USER', 'root'),
        password=os.getenv('DB_PASSWORD', ''),
        database=os.getenv('DB_NAME', 'inventario_repuestos'),
        cursorclass=pymysql.cursors.DictCursor
    )


def ejecutar_query(query, params=None, commit=False, fetch_one=False, fetch_all=False):
    """
    Función auxiliar para ejecutar consultas SQL de forma segura.
    """
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(query, params or ())
        
        if commit:
            connection.commit()
            last_id = cursor.lastrowid
            return last_id
        
        if fetch_one:
            result = cursor.fetchone()
            return result
        
        if fetch_all:
            results = cursor.fetchall()
            return results
        
        return None
        
    except Exception as e:
        print(f" Error en la base de datos: {e}")
        import logging
        logging.error(f"Error DB: {e}")
        if commit and connection:
            connection.rollback()
            return None
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()


# ---------------------------------------------------------------------------------
# CONSTANTES
# ---------------------------------------------------------------------------------
STOCK_MINIMO = 5

ROLES = {
    'admin': 'Administrador',
    'vendedor': 'Vendedor',
    'auditor': 'Auditor'
}


# ---------------------------------------------------------------------------------
# FUNCIONES DE CARGA Y GUARDADO (MYSQL)
# ---------------------------------------------------------------------------------
def cargar_productos():
    """Carga todos los productos desde MySQL"""
    query = """
        SELECT id, codigo_sku, nombre, categoria, marca, stock, precio_unitario, 
               porcentaje_ganancia, precio_venta, descripcion, valor_total, 
               fecha_creacion, fecha_actualizacion
        FROM productos
        ORDER BY id
    """
    productos = ejecutar_query(query, fetch_all=True)
    return productos if productos else []


def obtener_producto_por_id(producto_id):
    """Obtiene un producto específico por su ID"""
    query = """
        SELECT id, codigo_sku, nombre, categoria, marca, stock, precio_unitario, 
               porcentaje_ganancia, precio_venta, descripcion, valor_total, 
               fecha_creacion, fecha_actualizacion
        FROM productos
        WHERE id = %s
    """
    return ejecutar_query(query, (producto_id,), fetch_one=True)

def actualizar_producto(producto_id, nombre, categoria, marca, stock, precio_unitario, descripcion, codigo_sku=None, porcentaje_ganancia=0):
    """Actualiza un producto existente en MySQL"""
    valor_total = round(stock * precio_unitario, 3)
    
    precio_con_iva = round(precio_unitario * 1.19, 3)
    precio_venta = round(precio_con_iva * (1 + porcentaje_ganancia / 100), 3)

    if stock < 0 or precio_unitario < 0:
        flash("El stock y el precio no pueden ser negativos.", "error")
        return redirect(url_for('editar_producto', id=id))
    
    query = """
        UPDATE productos 
        SET nombre = %s, categoria = %s, marca = %s, stock = %s, 
            precio_unitario = %s, porcentaje_ganancia = %s, precio_venta = %s,
            descripcion = %s, valor_total = %s, codigo_sku = %s,
            fecha_actualizacion = NOW()
        WHERE id = %s
    """
    params = (nombre, categoria, marca, stock, precio_unitario, porcentaje_ganancia, 
              precio_venta, descripcion, valor_total, codigo_sku, producto_id)
    return ejecutar_query(query, params, commit=True)

def eliminar_producto_db(producto_id):
    """Elimina un producto de MySQL"""
    query = "DELETE FROM productos WHERE id = %s"
    return ejecutar_query(query, (producto_id,), commit=True)


def actualizar_stock_producto(producto_id, nuevo_stock):
    """Actualiza solo el stock de un producto"""
    producto = obtener_producto_por_id(producto_id)
    if not producto:
        return False
    
    valor_total = round(nuevo_stock * producto['precio_unitario'], 3)
    
    query = """
        UPDATE productos 
        SET stock = %s, valor_total = %s, fecha_actualizacion = NOW()
        WHERE id = %s
    """
    return ejecutar_query(query, (nuevo_stock, valor_total, producto_id), commit=True)

def reiniciar_autoincrement_productos():
    """Reinicia el AUTO_INCREMENT de la tabla productos si está vacía"""
    try:
        count_query = "SELECT COUNT(*) as total FROM productos"
        result = ejecutar_query(count_query, fetch_one=True)
        
        if result and result.get('total', 0) == 0:
            reset_query = "ALTER TABLE productos AUTO_INCREMENT = 1"
            ejecutar_query(reset_query, commit=True)
            return True
        return False
    except Exception as e:
        print(f"❌ Error al reiniciar AUTO_INCREMENT: {e}")
        return False


def reiniciar_autoincrement_ventas():
    """Reinicia el AUTO_INCREMENT de la tabla ventas si está vacía"""
    try:
        count_query = "SELECT COUNT(*) as total FROM ventas"
        result = ejecutar_query(count_query, fetch_one=True)
        
        if result and result.get('total', 0) == 0:
            reset_query = "ALTER TABLE ventas AUTO_INCREMENT = 1"
            ejecutar_query(reset_query, commit=True)
            return True
        return False
    except Exception as e:
        print(f"❌ Error al reiniciar AUTO_INCREMENT de ventas: {e}")
        return False


def cargar_ventas():
    """Carga todas las ventas desde MySQL"""
    query = """
        SELECT id, fecha, hora, producto_id, producto_nombre, categoria,
               cantidad, precio_unitario, iva_total, porcentaje_ganancia,
               ganancia_unitaria, ganancia_total, total,
               usuario_id, usuario_nombre, fecha_registro
        FROM ventas
        ORDER BY fecha DESC, hora DESC
    """
    ventas = ejecutar_query(query, fetch_all=True)
    return ventas if ventas else []

def guardar_venta(venta):
    """Guarda una nueva venta en MySQL"""
    query = """
        INSERT INTO ventas (
            fecha, hora, producto_id, producto_nombre, categoria,
            cantidad, precio_unitario, iva_total, porcentaje_ganancia,
            ganancia_unitaria, ganancia_total, total,
            usuario_id, usuario_nombre
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
    """
    params = (
        venta['fecha'],
        venta['hora'],
        venta['producto_id'],
        venta['producto_nombre'],
        venta['categoria'],
        venta['cantidad'],
        venta['precio_unitario'],
        venta.get('iva_total', 0),
        venta.get('porcentaje_ganancia', 0),
        venta.get('ganancia_unitaria', 0),
        venta.get('ganancia_total', 0),
        venta.get('total', 0),
        venta.get('usuario_id'),
        venta.get('usuario_nombre')
    )
    return ejecutar_query(query, params, commit=True)
    # ---------------------------------------------------------------------------------
# FUNCIONES DE USUARIOS Y AUTENTICACIÓN
# ---------------------------------------------------------------------------------
def cargar_usuarios():
    """Carga todos los usuarios desde MySQL"""
    query = """
        SELECT id, username, password, nombre_completo, rol, activo, 
               fecha_creacion, fecha_actualizacion
        FROM usuarios
        ORDER BY id
    """
    usuarios = ejecutar_query(query, fetch_all=True)
    
    if not usuarios:
        query_insert = """
            INSERT INTO usuarios (username, password, nombre_completo, rol, activo)
            VALUES (%s, %s, %s, %s, %s)
        """
        ejecutar_query(
            query_insert,
            ('admin', generate_password_hash('admin123'), 'Administrador', 'admin', True),
            commit=True
        )
        return cargar_usuarios()
    
    return usuarios


def crear_usuario(username, password, nombre_completo, rol):
    """Crea un nuevo usuario en MySQL"""
    query = """
        INSERT INTO usuarios (username, password, nombre_completo, rol, activo)
        VALUES (%s, %s, %s, %s, %s)
    """
    password_hash = generate_password_hash(password)
    params = (username, password_hash, nombre_completo, rol, True)
    return ejecutar_query(query, params, commit=True)


def actualizar_usuario_estado(usuario_id, activo):
    """Activa o desactiva un usuario"""
    query = """
        UPDATE usuarios 
        SET activo = %s, fecha_actualizacion = NOW()
        WHERE id = %s
    """
    return ejecutar_query(query, (activo, usuario_id), commit=True)


def eliminar_usuario_db(usuario_id):
    """Elimina un usuario de MySQL"""
    query = "DELETE FROM usuarios WHERE id = %s"
    return ejecutar_query(query, (usuario_id,), commit=True)


def actualizar_password_usuario(usuario_id, nueva_password):
    """Actualiza la contraseña de un usuario"""
    query = """
        UPDATE usuarios 
        SET password = %s, fecha_actualizacion = NOW()
        WHERE id = %s
    """
    password_hash = generate_password_hash(nueva_password)
    return ejecutar_query(query, (password_hash, usuario_id), commit=True)


def obtener_usuario_por_id(usuario_id):
    """Obtiene un usuario específico por su ID"""
    query = """
        SELECT id, username, password, nombre_completo, rol, activo, 
               fecha_creacion, fecha_actualizacion
        FROM usuarios
        WHERE id = %s
    """
    return ejecutar_query(query, (usuario_id,), fetch_one=True)


def registrar_log(accion, detalle=""):
    """Registra una acción en los logs de MySQL"""
    query = """
        INSERT INTO logs (fecha, hora, usuario, accion, detalle)
        VALUES (%s, %s, %s, %s, %s)
    """
    params = (
        datetime.now().strftime('%Y-%m-%d'),
        datetime.now().strftime('%H:%M:%S'),
        session.get('username', 'Sistema'),
        accion,
        detalle
    )
    ejecutar_query(query, params, commit=True)


def cargar_logs():
    """Carga todos los logs desde MySQL"""
    query = """
    SELECT id, fecha, hora, usuario, accion, detalle, fecha_registro
    FROM logs
    ORDER BY fecha_registro ASC
    """
    logs = ejecutar_query(query, fetch_all=True)
    return logs if logs else []


# ---------------------------------------------------------------------------------
# DECORADORES DE AUTENTICACIÓN
# ---------------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Debes iniciar sesión para acceder a esta página.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Debes iniciar sesión.', 'warning')
                return redirect(url_for('login'))
            
            if session.get('rol') not in roles:
                flash('No tienes permisos para acceder a esta página.', 'danger')
                return redirect(url_for('inicio'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ---------------------------------------------------------------------------------
# RUTAS DE AUTENTICACIÓN
# ---------------------------------------------------------------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de inicio de sesión"""
    if 'user_id' in session:
        return redirect(url_for('inicio'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('Por favor completa todos los campos.', 'error')
            return redirect(url_for('login'))
        
        usuarios = cargar_usuarios()
        usuario = next((u for u in usuarios if u['username'] == username), None)
        
        if not usuario:
            flash('Usuario no encontrado.', 'error')
            return redirect(url_for('login'))
        
        if not usuario.get('activo', True):
            flash('Tu cuenta está desactivada. Contacta al administrador.', 'error')
            return redirect(url_for('login'))
        
        if not check_password_hash(usuario['password'], password):
            flash('Contraseña incorrecta.', 'error')
            return redirect(url_for('login'))
        
        session['user_id'] = usuario['id']
        session['username'] = usuario['username']
        session['nombre_completo'] = usuario['nombre_completo']
        session['rol'] = usuario['rol']
        
        registrar_log('Inicio de sesión', f"Usuario: {usuario['username']}")
        
        flash(f'¡Bienvenido {usuario["nombre_completo"]}!', 'success')
        return redirect(url_for('inicio'))
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    """Cerrar sesión"""
    username = session.get('username', 'Usuario')
    registrar_log('Cierre de sesión', f"Usuario: {username}")
    
    session.clear()
    flash('Has cerrado sesión correctamente.', 'info')
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------------
# RUTAS DE GESTIÓN DE USUARIOS
# ---------------------------------------------------------------------------------
@app.route('/usuarios')
@login_required
@role_required('admin')
def lista_usuarios():
    """Lista de usuarios (solo admin)"""
    usuarios = cargar_usuarios()
    return render_template('usuarios/lista_usuarios.html', usuarios=usuarios, roles=ROLES)


@app.route('/usuarios/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def nuevo_usuario():
    """Crear nuevo usuario (solo admin)"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        nombre_completo = request.form.get('nombre_completo', '').strip()
        rol = request.form.get('rol', '')
        
        if not username or not password or not nombre_completo or not rol:
            flash('Todos los campos son obligatorios.', 'error')
            return redirect(url_for('nuevo_usuario'))
        
        if rol not in ROLES:
            flash('Rol inválido.', 'error')
            return redirect(url_for('nuevo_usuario'))
        
        usuarios = cargar_usuarios()
        
        if any(u['username'] == username for u in usuarios):
            flash(f'El usuario "{username}" ya existe.', 'error')
            return redirect(url_for('nuevo_usuario'))
        
        usuario_id = crear_usuario(username, password, nombre_completo, rol)
        
        if usuario_id:
            registrar_log('Usuario creado', f"Nuevo usuario: {username} ({ROLES[rol]})")
            flash(f'Usuario "{username}" creado exitosamente.', 'success')
            return redirect(url_for('lista_usuarios'))
        else:
            flash('Error al crear el usuario.', 'error')
            return redirect(url_for('nuevo_usuario'))
    
    return render_template('usuarios/nuevo_usuario.html', roles=ROLES)


@app.route('/usuarios/<int:id>/toggle')
@login_required
@role_required('admin')
def toggle_usuario(id):
    """Activar/desactivar usuario"""
    usuario = obtener_usuario_por_id(id)
    
    if not usuario:
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('lista_usuarios'))
    
    if usuario['id'] == session.get('user_id'):
        flash('No puedes desactivar tu propia cuenta.', 'error')
        return redirect(url_for('lista_usuarios'))
    
    nuevo_estado = not usuario.get('activo', True)
    actualizar_usuario_estado(id, nuevo_estado)
    
    estado = "activado" if nuevo_estado else "desactivado"
    registrar_log(f'Usuario {estado}', f"Usuario: {usuario['username']}")
    
    flash(f'Usuario "{usuario["username"]}" {estado}.', 'success')
    return redirect(url_for('lista_usuarios'))


@app.route('/usuarios/<int:id>/eliminar')
@login_required
@role_required('admin')
def eliminar_usuario(id):
    """Eliminar usuario (solo admin)"""
    usuario = obtener_usuario_por_id(id)
    
    if not usuario:
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('lista_usuarios'))
    
    if usuario['id'] == session.get('user_id'):
        flash('No puedes eliminar tu propia cuenta.', 'error')
        return redirect(url_for('lista_usuarios'))
    
    if usuario['username'] == 'admin':
        flash('No se puede eliminar el usuario administrador principal.', 'error')
        return redirect(url_for('lista_usuarios'))
    
    eliminar_usuario_db(id)
    
    registrar_log('Usuario eliminado', f"Usuario: {usuario['username']}")
    
    flash(f'Usuario "{usuario["username"]}" eliminado.', 'success')
    return redirect(url_for('lista_usuarios'))


# ---------------------------------------------------------------------------------
# GESTIÓN DE CONTRASEÑAS
# ---------------------------------------------------------------------------------
@app.route('/cambiar-password', methods=['GET', 'POST'])
@login_required
def cambiar_password():
    """Permite al usuario cambiar su propia contraseña"""
    if request.method == 'POST':
        password_actual = request.form.get('password_actual', '')
        password_nueva = request.form.get('password_nueva', '')
        password_confirmar = request.form.get('password_confirmar', '')
        
        if not password_actual or not password_nueva or not password_confirmar:
            flash('Todos los campos son obligatorios.', 'error')
            return redirect(url_for('cambiar_password'))
        
        if password_nueva != password_confirmar:
            flash('Las contraseñas nuevas no coinciden.', 'error')
            return redirect(url_for('cambiar_password'))
        
        if len(password_nueva) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.', 'error')
            return redirect(url_for('cambiar_password'))
        
        usuario = obtener_usuario_por_id(session['user_id'])
        
        if not usuario:
            flash('Usuario no encontrado.', 'error')
            return redirect(url_for('cambiar_password'))
        
        if not check_password_hash(usuario['password'], password_actual):
            flash('La contraseña actual es incorrecta.', 'error')
            return redirect(url_for('cambiar_password'))
        
        actualizar_password_usuario(usuario['id'], password_nueva)
        
        registrar_log('Contraseña cambiada', f"Usuario: {usuario['username']}")
        
        flash('¡Contraseña actualizada exitosamente!', 'success')
        return redirect(url_for('inicio'))
    
    return render_template('cambiar_password.html')


@app.route('/usuarios/<int:id>/resetear-password', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def resetear_password_usuario(id):
    """Admin puede resetear la contraseña de un usuario"""
    usuario = obtener_usuario_por_id(id)
    
    if not usuario:
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('lista_usuarios'))
    
    if request.method == 'POST':
        nueva_password = request.form.get('nueva_password', '')
        confirmar_password = request.form.get('confirmar_password', '')
        
        if not nueva_password or not confirmar_password:
            flash('Todos los campos son obligatorios.', 'error')
            return redirect(url_for('resetear_password_usuario', id=id))
        
        if nueva_password != confirmar_password:
            flash('Las contraseñas no coinciden.', 'error')
            return redirect(url_for('resetear_password_usuario', id=id))
        
        if len(nueva_password) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.', 'error')
            return redirect(url_for('resetear_password_usuario', id=id))
        
        actualizar_password_usuario(id, nueva_password)
        
        registrar_log('Contraseña reseteada por admin', f"Admin reseteó contraseña de: {usuario['username']}")
        
        flash(f'Contraseña de "{usuario["username"]}" actualizada exitosamente.', 'success')
        return redirect(url_for('lista_usuarios'))
    
    return render_template('usuarios/resetear_password.html', usuario=usuario)


@app.route('/logs')
@login_required
@role_required('admin', 'auditor')
def ver_logs():
    """Ver logs del sistema"""
    logs = cargar_logs()
    return render_template('logs.html', logs=logs)


@app.route('/logs/eliminar/<int:id>', methods=['POST'])
@login_required
@role_required('admin')
def eliminar_log(id):
    """Elimina un log individual"""
    query = "DELETE FROM logs WHERE id = %s"
    ejecutar_query(query, (id,), commit=True)

    registrar_log("Log eliminado", f"ID del log eliminado: {id}")
    flash("Registro eliminado correctamente.", "success")
    return redirect(url_for('ver_logs'))


@app.route('/logs/limpiar', methods=['POST'])
@login_required
@role_required('admin')
def limpiar_logs():
    """Elimina todos los registros del historial y reinicia el AUTO_INCREMENT"""
    ejecutar_query("DELETE FROM logs", commit=True)
    ejecutar_query("ALTER TABLE logs AUTO_INCREMENT = 1", commit=True)
    registrar_log("Historial limpiado", "Se eliminaron todos los logs y se reinició el AUTO_INCREMENT.")
    flash("Historial limpiado correctamente. IDs reiniciados desde 1.", "success")
    return redirect(url_for('ver_logs'))


# ---------------------------------------------------------------------------------
# RUTAS PRINCIPALES DE PRODUCTOS
# ---------------------------------------------------------------------------------
@app.route('/')
@login_required
def inicio():
    """Página de inicio - redirige a lista de productos"""
    return redirect(url_for('lista_productos'))


@app.route('/productos', methods=['GET'])
@login_required
def lista_productos():
    """Lista todos los productos con filtros y búsqueda"""
    productos = cargar_productos()

    query = request.args.get("q", "").lower().strip()
    categoria = request.args.get("categoria", "").strip()
    ordenar = request.args.get("ordenar", "")
    solo_bajo_stock = request.args.get("solo_bajo_stock", "0") == "1"

    if query:
        productos = [p for p in productos if query in p.get('nombre', '').lower()]

    if categoria:
        productos = [p for p in productos if p.get('categoria') == categoria]

    if solo_bajo_stock:
        productos = [p for p in productos if p.get('stock', 0) < STOCK_MINIMO]

    if ordenar == "precio_asc":
        productos = sorted(productos, key=lambda x: x.get("precio_unitario", 0))
    elif ordenar == "precio_desc":
        productos = sorted(productos, key=lambda x: x.get("precio_unitario", 0), reverse=True)
    elif ordenar == "stock_asc":
        productos = sorted(productos, key=lambda x: x.get("stock", 0))
    elif ordenar == "stock_desc":
        productos = sorted(productos, key=lambda x: x.get("stock", 0), reverse=True)

    todos_productos = cargar_productos()
    bajo_stock_total = sum(1 for p in todos_productos if p.get('stock', 0) < STOCK_MINIMO)
    categorias = sorted({p.get("categoria", "") for p in todos_productos})

    return render_template(
        "productos.html",
        productos=productos,
        categorias=categorias,
        bajo_stock_total=bajo_stock_total,
        solo_bajo_stock=solo_bajo_stock,
        STOCK_MINIMO=STOCK_MINIMO
    )


@app.route('/productos/nuevo', methods=["GET", "POST"])
@login_required
@role_required('admin', 'vendedor')
def nuevo_producto():
    """Crea un nuevo producto"""
    if request.method == "POST":
        try:
            nombre = request.form['nombre'].strip()
            categoria = request.form['categoria'].strip()
            marca = request.form.get('marca', '').strip()
            stock = int(request.form['stock'])
            precio_unitario = round(float(request.form['precio_unitario']), 3)
            descripcion = request.form.get('descripcion', '').strip()
            codigo_sku = request.form.get('codigo_sku', '').strip()
            porcentaje_ganancia = float(request.form.get('porcentaje_ganancia', 0))

            if not nombre or not categoria:
                flash("El nombre y la categoría son obligatorios.", "error")
                return redirect(url_for('nuevo_producto'))

            if stock < 0 or precio_unitario < 0:
                flash("El stock y el precio no pueden ser negativos.", "error")
                return redirect(url_for('nuevo_producto'))

        except ValueError:
            flash("Error en los datos del formulario. Verifica los valores numéricos.", "error")
            return redirect(url_for('nuevo_producto'))
        except KeyError:
            flash("Faltan campos obligatorios en el formulario.", "error")
            return redirect(url_for('nuevo_producto'))

        precio_con_iva = round(precio_unitario * 1.19, 3)
        precio_venta = round(precio_con_iva * (1 + porcentaje_ganancia / 100), 3)

        query = """
            INSERT INTO productos (nombre, categoria, marca, stock, precio_unitario, 
                                   porcentaje_ganancia, precio_venta, descripcion, codigo_sku)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        params = (nombre, categoria, marca, stock, precio_unitario, 
                  porcentaje_ganancia, precio_venta, descripcion, codigo_sku)
        producto_id = ejecutar_query(query, params, commit=True)

        if producto_id:
            registrar_log('Producto creado', f"Producto: {nombre} (ID: {producto_id})")
            flash(f"El producto '{nombre}' ha sido registrado con éxito.", "success")
            return redirect(url_for('lista_productos'))
        else:
            flash("Error al guardar el producto en la base de datos.", "error")
            return redirect(url_for('nuevo_producto'))

    return render_template("crear_producto.html")


@app.route('/productos/<int:id>', methods=["GET"])
@login_required
def detalle_producto(id):
    """Muestra los detalles de un producto"""
    producto = obtener_producto_por_id(id)

    if not producto:
        flash(f"El producto con el ID {id} no fue encontrado.", "error")
        return redirect(url_for('lista_productos'))

    return render_template("detalle_producto.html", producto=producto)


@app.route('/productos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'vendedor')
def editar_producto(id):
    """Edita un producto existente"""
    producto = obtener_producto_por_id(id)

    if not producto:
        flash(f"El producto con el ID {id} no pudo ser encontrado.", "error")
        return redirect(url_for('lista_productos'))

    if request.method == 'POST':
        try:
            nombre = request.form['nombre'].strip()
            categoria = request.form['categoria'].strip()
            marca = request.form.get('marca', '').strip()
            stock = int(request.form['stock'])
            precio_unitario = round(float(request.form['precio_unitario']), 3)
            descripcion = request.form.get('descripcion', '').strip()
            codigo_sku = request.form.get('codigo_sku', '').strip()
            porcentaje_ganancia = float(request.form.get('porcentaje_ganancia', 0))

            if stock < 0 or precio_unitario < 0:
                flash("El stock y el precio no pueden ser negativos.", "error")
                return redirect(url_for('editar_producto', id=id))

            actualizar_producto(id, nombre, categoria, marca, stock, precio_unitario, 
                              descripcion, codigo_sku, porcentaje_ganancia)

            registrar_log('Producto actualizado', f"Producto: {nombre} (ID: {id})")

            flash(f"El producto '{nombre}' fue actualizado con éxito.", "success")
            return redirect(url_for('lista_productos'))

        except ValueError:
            flash("Error en los datos del formulario.", "error")
            return redirect(url_for('editar_producto', id=id))

    return render_template("editar_producto.html", producto=producto)


@app.route("/productos/eliminar/<int:id>", methods=['POST'])
@login_required
@role_required('admin')
def eliminar_producto(id):
    """Elimina un producto - VERSIÓN MEJORADA CON MANEJO DE VENTAS Y RESET AUTOMÁTICO"""
    
    try:
        producto = obtener_producto_por_id(id)
        
        if not producto:
            return jsonify({
                'success': False, 
                'error': 'Producto no encontrado'
            }), 404
        
        query_ventas = "SELECT COUNT(*) as total FROM ventas WHERE producto_id = %s"
        resultado = ejecutar_query(query_ventas, (id,), fetch_one=True)
        
        ventas_asociadas = resultado['total'] if resultado else 0
        
        if ventas_asociadas > 0:
            return jsonify({
                'success': False,
                'error': f'No se puede eliminar. Tiene {ventas_asociadas} venta(s) asociada(s)',
                'ventas_asociadas': ventas_asociadas
            }), 400
        
        query_eliminar = "DELETE FROM productos WHERE id = %s"
        ejecutar_query(query_eliminar, (id,), commit=True)
        
        registrar_log('Producto eliminado', f"Producto: {producto['nombre']} (ID: {id})")
        
        count_query = "SELECT COUNT(*) as total FROM productos"
        count_result = ejecutar_query(count_query, fetch_one=True)
        
        mensaje = 'Producto eliminado exitosamente'
        
        if count_result and count_result.get('total', 0) == 0:
            reset_query = "ALTER TABLE productos AUTO_INCREMENT = 1"
            ejecutar_query(reset_query, commit=True)
            registrar_log('IDs reseteados', 'AUTO_INCREMENT reiniciado automáticamente')
            mensaje = 'Producto eliminado. IDs reseteados a 1 (no quedan productos)'
        
        return jsonify({
            'success': True,
            'message': mensaje
        }), 200
        
    except Exception as e:
        print(f"❌ Error al eliminar producto: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ---------------------------------------------------------------------------------
# RUTAS DE VENTAS (🔥 CORREGIDO)
# ---------------------------------------------------------------------------------
@app.route('/ventas/nueva', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'vendedor')
def nueva_venta():
    """Registra una nueva venta con cálculo automático de ganancias"""
    productos = cargar_productos()

    if request.method == "POST":
        try:
            # Obtener datos del formulario
            producto_id = int(request.form['producto_id'])
            cantidad = int(request.form['cantidad'])
            ganancia_general = float(request.form.get('porcentaje_ganancia_general', 0))

            # Validar cantidad
            if cantidad <= 0:
                flash("La cantidad debe ser mayor a 0.", "error")
                return redirect(url_for('nueva_venta'))

        except (ValueError, KeyError) as e:
            flash(f"Datos inválidos en el formulario: {str(e)}", "error")
            return redirect(url_for('nueva_venta'))

        # Obtener producto de la base de datos
        producto = obtener_producto_por_id(producto_id)

        if not producto:
            flash("Producto no encontrado.", "error")
            return redirect(url_for('nueva_venta'))

        # Validar stock disponible
        if cantidad > producto['stock']:
            flash(f"Stock insuficiente. Disponible: {producto['stock']} unidades.", "error")
            return redirect(url_for('nueva_venta'))

        # ========================================
        # CÁLCULOS FINANCIEROS
        # ========================================
        
        # 1. Precio base (sin IVA)
        precio_base = float(producto['precio_unitario'])
        
        # 2. Calcular IVA (19%)
        iva_unitario = round(precio_base * 0.19, 3)
        precio_con_iva = round(precio_base + iva_unitario, 3)
        iva_total = round(iva_unitario * cantidad, 3)
        
        # 3. Determinar porcentaje de ganancia
        if ganancia_general > 0:
            # Usar ganancia general ingresada
            porcentaje_ganancia = ganancia_general
        else:
            # Usar ganancia del producto (o 0 si no tiene)
            porcentaje_ganancia = float(producto.get('porcentaje_ganancia', 0))
        
        # 4. Calcular precio de venta
        if porcentaje_ganancia > 0:
            precio_venta_unitario = round(precio_con_iva * (1 + porcentaje_ganancia / 100), 3)
        else:
            precio_venta_unitario = precio_con_iva
        
        # 5. Calcular ganancias
        ganancia_unitaria = round(precio_venta_unitario - precio_con_iva, 3)
        ganancia_total = round(ganancia_unitaria * cantidad, 3)
        
        # 6. Calcular total de la venta
        total_venta = round(precio_venta_unitario * cantidad, 3)

        # ========================================
        # CREAR REGISTRO DE VENTA
        # ========================================
        venta = {
            'fecha': datetime.now().strftime('%Y-%m-%d'),
            'hora': datetime.now().strftime('%H:%M:%S'),
            'producto_id': producto['id'],
            'producto_nombre': producto['nombre'],
            'categoria': producto['categoria'],
            'cantidad': cantidad,
            'precio_unitario': precio_base,
            'iva_total': iva_total,
            'porcentaje_ganancia': porcentaje_ganancia,
            'ganancia_unitaria': ganancia_unitaria,
            'ganancia_total': ganancia_total,
            'total': total_venta,
            'usuario_id': session.get('user_id'),
            'usuario_nombre': session.get('nombre_completo')
        }
        
        # Guardar en base de datos
        venta_id = guardar_venta(venta)

        if venta_id:
            # Registrar en logs
            registrar_log(
                'Venta registrada', 
                f"{cantidad} x {producto['nombre']} | Total: ${total_venta:,.0f} | Ganancia: ${ganancia_total:,.0f} ({porcentaje_ganancia}%)"
            )

            # Actualizar stock del producto
            nuevo_stock = producto['stock'] - cantidad
            actualizar_stock_producto(producto_id, nuevo_stock)

            # Mostrar mensaje de éxito
            flash(
                f"✅ Venta registrada exitosamente\n"
                f"Producto: {producto['nombre']} x{cantidad}\n"
                f"Total: ${total_venta:,.0f} COP | Ganancia: ${ganancia_total:,.0f} COP ({porcentaje_ganancia:.1f}%)", 
                "success"
            )
            return redirect(url_for('historial_ventas'))
        else:
            flash("❌ Error al guardar la venta en la base de datos.", "error")
            return redirect(url_for('nueva_venta'))

    return render_template("crear_venta.html", productos=productos)


# 🔥 FUNCIÓN CORREGIDA DEL HISTORIAL DE VENTAS
@app.route('/ventas/historial')
@login_required
def historial_ventas():
    """Muestra el historial de ventas con filtros"""
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    
    # ✅ SOLO seleccionar columnas que EXISTEN en la tabla
    query = """
        SELECT 
            v.id,
            v.fecha,
            v.hora,
            v.producto_id,
            v.producto_nombre,
            v.categoria,
            v.cantidad,
            v.precio_unitario,
            v.iva_total,
            v.porcentaje_ganancia,
            v.ganancia_unitaria,
            v.ganancia_total,
            v.total,
            v.usuario_id,
            v.usuario_nombre,
            v.fecha_registro
        FROM ventas v
        WHERE 1=1
    """
    
    params = []
    
    # Aplicar filtros si existen
    if fecha_desde:
        query += " AND v.fecha >= %s"
        params.append(fecha_desde)
    
    if fecha_hasta:
        query += " AND v.fecha <= %s"
        params.append(fecha_hasta)
    
    query += " ORDER BY v.id ASC"
    
    # Ejecutar query
    ventas = ejecutar_query(query, tuple(params) if params else None, fetch_all=True)

    # Procesar ventas y asegurar valores por defecto
    ventas_procesadas = []
    if ventas:
        for v in ventas:
            venta_dict = dict(v) if isinstance(v, dict) else {
                'id': v[0],
                'fecha': v[1],
                'hora': v[2],
                'producto_id': v[3],
                'producto_nombre': v[4],
                'categoria': v[5],
                'cantidad': v[6],
                'precio_unitario': v[7],
                'iva_total': v[8],
                'porcentaje_ganancia': v[9],
                'ganancia_unitaria': v[10],
                'ganancia_total': v[11],
                'total': v[12],
                'usuario_id': v[13],
                'usuario_nombre': v[14],
                'fecha_registro': v[15]
            }
            
            # Asegurar que campos numéricos tengan valores por defecto
            venta_dict['iva_total'] = venta_dict.get('iva_total') or 0
            venta_dict['porcentaje_ganancia'] = venta_dict.get('porcentaje_ganancia') or 0
            venta_dict['ganancia_unitaria'] = venta_dict.get('ganancia_unitaria') or 0
            venta_dict['ganancia_total'] = venta_dict.get('ganancia_total') or 0
            
            ventas_procesadas.append(venta_dict)
    
    # Calcular totales generales
    total_vendido = sum(float(v.get('total', 0)) for v in ventas_procesadas)
    total_iva = sum(float(v.get('iva_total', 0)) for v in ventas_procesadas)
    total_ganancias = sum(float(v.get('ganancia_total', 0)) for v in ventas_procesadas)
    num_ventas = len(ventas_procesadas)

    return render_template(
        'historial_ventas.html',
        ventas=ventas_procesadas,
        total_vendido=total_vendido,
        total_general=total_vendido,  # Alias por compatibilidad
        total_iva=total_iva,
        total_ganancias=total_ganancias,
        num_ventas=num_ventas
    )


@app.route('/ventas/eliminar/<int:id>', methods=['POST'])
@login_required
@role_required('admin')
def eliminar_venta(id):
    """Elimina una venta del historial"""
    try:
        query = "DELETE FROM ventas WHERE id = %s"
        ejecutar_query(query, (id,), commit=True)
        
        registrar_log('Venta eliminada', f'ID de venta eliminada: {id}')
        
        if reiniciar_autoincrement_ventas():
            flash('Venta eliminada. IDs reiniciados a 1 (no quedan ventas).', 'success')
        else:
            flash('Venta eliminada correctamente.', 'success')
            
    except Exception as e:
        flash(f'Error al eliminar la venta: {str(e)}', 'error')
    
    return redirect(url_for('historial_ventas'))


@app.route('/ventas/<int:id>/detalle')
@login_required
def detalle_venta(id):
    """Muestra los detalles de una venta específica"""
    query = """
        SELECT 
            v.id, v.fecha, v.hora, v.producto_id, v.producto_nombre,
            v.categoria, v.cantidad, v.precio_unitario, v.iva_total,
            v.porcentaje_ganancia, v.ganancia_unitaria, v.ganancia_total,
            v.total, v.usuario_id, v.usuario_nombre, v.fecha_registro
        FROM ventas v
        WHERE v.id = %s
    """
    venta = ejecutar_query(query, (id,), fetch_one=True)
    
    if not venta:
        flash('Venta no encontrada.', 'error')
        return redirect(url_for('historial_ventas'))
    
    venta_dict = dict(venta)
    venta_dict['iva_total'] = venta_dict.get('iva_total') or 0
    venta_dict['porcentaje_ganancia'] = venta_dict.get('porcentaje_ganancia') or 0
    venta_dict['ganancia_unitaria'] = venta_dict.get('ganancia_unitaria') or 0
    venta_dict['ganancia_total'] = venta_dict.get('ganancia_total') or 0
    
    return render_template('detalle_venta.html', venta=venta_dict)


@app.route('/ventas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'vendedor')
def editar_venta(id):
    """Edita una venta existente"""
    if request.method == 'GET':
        query = """
            SELECT 
                v.id, v.fecha, v.hora, v.producto_id, v.producto_nombre,
                v.categoria, v.cantidad, v.precio_unitario, v.iva_total,
                v.porcentaje_ganancia, v.ganancia_unitaria, v.ganancia_total,
                v.total, v.usuario_id, v.usuario_nombre
            FROM ventas v
            WHERE v.id = %s
        """
        venta = ejecutar_query(query, (id,), fetch_one=True)
        
        if not venta:
            flash('Venta no encontrada.', 'error')
            return redirect(url_for('historial_ventas'))
        
        productos = cargar_productos()
        return render_template('editar_venta.html', venta=dict(venta), productos=productos)
    
    # POST - Actualizar venta
    try:
        producto_id = int(request.form['producto_id'])
        cantidad = int(request.form['cantidad'])
        ganancia_general = float(request.form.get('porcentaje_ganancia_general', 0))
        
        if cantidad <= 0:
            flash("La cantidad debe ser mayor a 0.", "error")
            return redirect(url_for('editar_venta', id=id))
    except (ValueError, KeyError):
        flash("Datos inválidos.", "error")
        return redirect(url_for('editar_venta', id=id))
    
    venta_actual = ejecutar_query(
        "SELECT producto_id, cantidad FROM ventas WHERE id = %s",
        (id,),
        fetch_one=True
    )
    
    if not venta_actual:
        flash("Venta no encontrada.", "error")
        return redirect(url_for('historial_ventas'))
    
    producto = obtener_producto_por_id(producto_id)
    
    if not producto:
        flash("Producto no encontrado.", "error")
        return redirect(url_for('editar_venta', id=id))
    
    diferencia_stock = venta_actual['cantidad'] - cantidad
    nuevo_stock = producto['stock'] + diferencia_stock
    
    if nuevo_stock < 0:
        flash(f"Stock insuficiente. Disponible: {producto['stock']} unidades.", "error")
        return redirect(url_for('editar_venta', id=id))
    
    precio_base = float(producto['precio_unitario'])
    iva_unitario = round(precio_base * 0.19, 3)
    precio_con_iva = round(precio_base + iva_unitario, 3)
    iva_total = round(iva_unitario * cantidad, 3)
    
    porcentaje_ganancia = ganancia_general if ganancia_general > 0 else float(producto.get('porcentaje_ganancia', 0))
    
    if porcentaje_ganancia > 0:
        precio_venta_unitario = round(precio_con_iva * (1 + porcentaje_ganancia / 100), 3)
    else:
        precio_venta_unitario = precio_con_iva
    
    ganancia_unitaria = round(precio_venta_unitario - precio_con_iva, 3)
    ganancia_total = round(ganancia_unitaria * cantidad, 3)
    total_venta = round(precio_venta_unitario * cantidad, 3)
    
    query_update = """
        UPDATE ventas SET
            producto_id = %s,
            producto_nombre = %s,
            categoria = %s,
            cantidad = %s,
            precio_unitario = %s,
            iva_total = %s,
            porcentaje_ganancia = %s,
            ganancia_unitaria = %s,
            ganancia_total = %s,
            total = %s
        WHERE id = %s
    """
    
    ejecutar_query(
        query_update,
        (producto_id, producto['nombre'], producto['categoria'], 
         cantidad, precio_base, iva_total, porcentaje_ganancia, 
         ganancia_unitaria, ganancia_total, total_venta, id),
        commit=True
    )
    
    actualizar_stock_producto(producto_id, nuevo_stock)
    
    if venta_actual['producto_id'] != producto_id:
        producto_anterior = obtener_producto_por_id(venta_actual['producto_id'])
        if producto_anterior:
            stock_anterior_nuevo = producto_anterior['stock'] + venta_actual['cantidad']
            actualizar_stock_producto(venta_actual['producto_id'], stock_anterior_nuevo)
    
    registrar_log('Venta editada', f"ID: {id} | Producto: {producto['nombre']} x{cantidad}")
    
    flash(f"Venta #{id} actualizada exitosamente.", "success")
    return redirect(url_for('historial_ventas'))



# ---------------------------------------------------------------------------------
# DASHBOARD Y REPORTES
# ---------------------------------------------------------------------------------
@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard con métricas y estadísticas generales"""
    productos = cargar_productos()
    ventas = cargar_ventas()

    total_productos = len(productos)
    valor_inventario_total = sum(p.get('valor_total', 0) for p in productos)
    productos_bajo_stock = sum(1 for p in productos if p.get('stock', 0) < STOCK_MINIMO)
    total_ventas_realizadas = len(ventas)
    ingresos_totales = sum(v.get('total', 0) for v in ventas)

    ventas_por_producto = {}
    for v in ventas:
        pid = v.get('producto_id')
        if pid not in ventas_por_producto:
            ventas_por_producto[pid] = {
                'nombre': v.get('producto_nombre'),
                'cantidad_vendida': 0,
                'ingresos': 0
            }
        ventas_por_producto[pid]['cantidad_vendida'] += v.get('cantidad', 0)
        ventas_por_producto[pid]['ingresos'] += v.get('total', 0)

    top_vendidos = sorted(
        ventas_por_producto.values(),
        key=lambda x: x['cantidad_vendida'],
        reverse=True
    )[:5]

    ventas_por_categoria = {}
    for v in ventas:
        cat = v.get('categoria', 'Sin categoría')
        if cat not in ventas_por_categoria:
            ventas_por_categoria[cat] = {
                'cantidad': 0,
                'ingresos': 0
            }
        ventas_por_categoria[cat]['cantidad'] += v.get('cantidad', 0)
        ventas_por_categoria[cat]['ingresos'] += v.get('total', 0)

    ventas_por_dia = defaultdict(lambda: {'cantidad': 0, 'ingresos': 0})

    for v in ventas:
        fecha = v.get('fecha')
        ventas_por_dia[fecha]['cantidad'] += v.get('cantidad', 0)
        ventas_por_dia[fecha]['ingresos'] += v.get('total', 0)

    ventas_diarias = sorted(
        [{'fecha': k, 'cantidad': v['cantidad'], 'ingresos': v['ingresos']}
         for k, v in ventas_por_dia.items() if k],
        key=lambda x: x['fecha']
    )[-7:]

    productos_criticos = [
        p for p in productos
        if p.get('stock', 0) < STOCK_MINIMO
    ]
    productos_criticos = sorted(productos_criticos, key=lambda x: x.get('stock', 0))[:10]

    return render_template(
        'dashboard.html',
        total_productos=total_productos,
        valor_inventario_total=valor_inventario_total,
        productos_bajo_stock=productos_bajo_stock,
        total_ventas_realizadas=total_ventas_realizadas,
        ingresos_totales=ingresos_totales,
        top_vendidos=top_vendidos,
        ventas_por_categoria=ventas_por_categoria,
        ventas_diarias=ventas_diarias,
        productos_criticos=productos_criticos,
        STOCK_MINIMO=STOCK_MINIMO
    )


@app.route('/reportes/productos-mas-vendidos')
@login_required
def reporte_mas_vendidos():
    """Reporte detallado de productos más vendidos"""
    ventas = cargar_ventas()

    if not ventas:
        flash("No hay ventas registradas para generar el reporte.", "info")
        return redirect(url_for('dashboard'))

    ventas_por_producto = {}
    for v in ventas:
        pid = v.get('producto_id')
        if pid not in ventas_por_producto:
            ventas_por_producto[pid] = {
                'id': pid,
                'nombre': v.get('producto_nombre'),
                'categoria': v.get('categoria'),
                'cantidad_vendida': 0,
                'ingresos': 0,
                'num_ventas': 0
            }
        ventas_por_producto[pid]['cantidad_vendida'] += v.get('cantidad', 0)
        ventas_por_producto[pid]['ingresos'] += v.get('total', 0)
        ventas_por_producto[pid]['num_ventas'] += 1

    productos_vendidos = list(ventas_por_producto.values())

    top_5_mas = sorted(
        productos_vendidos,
        key=lambda x: x['cantidad_vendida'],
        reverse=True
    )[:5]

    top_5_menos = sorted(
        productos_vendidos,
        key=lambda x: x['cantidad_vendida']
    )[:5]

    todos_ordenados = sorted(
        productos_vendidos,
        key=lambda x: x['cantidad_vendida'],
        reverse=True
    )

    return render_template(
        'reporte_productos.html',
        top_5_mas=top_5_mas,
        top_5_menos=top_5_menos,
        todos_productos=todos_ordenados,
        total_productos=len(productos_vendidos)
    )


@app.route('/reportes/ventas-por-periodo')
@login_required
def reporte_ventas_periodo():
    """Reporte de ventas por período usando MySQL"""

    ventas = cargar_ventas()

    if not ventas:
        flash("No hay ventas registradas.", "info")
        return redirect(url_for('dashboard'))

    ventas_diarias = defaultdict(lambda: {'cantidad': 0, 'ingresos': 0, 'num_ventas': 0})
    ventas_mensuales = defaultdict(lambda: {'cantidad': 0, 'ingresos': 0, 'num_ventas': 0})

    for v in ventas:
        fecha = v.get('fecha')

        if isinstance(fecha, datetime):
            fecha_str = fecha.strftime("%Y-%m-%d")
        elif hasattr(fecha, "strftime"):  
            fecha_str = fecha.strftime("%Y-%m-%d")
        else:
            fecha_str = str(fecha)

        ventas_diarias[fecha_str]['cantidad'] += v.get('cantidad', 0)
        ventas_diarias[fecha_str]['ingresos'] += v.get('total', 0)
        ventas_diarias[fecha_str]['num_ventas'] += 1

        mes = fecha_str[:7]
        ventas_mensuales[mes]['cantidad'] += v.get('cantidad', 0)
        ventas_mensuales[mes]['ingresos'] += v.get('total', 0)
        ventas_mensuales[mes]['num_ventas'] += 1

    ventas_por_dia = sorted(
        [{'fecha': k, **v} for k, v in ventas_diarias.items()],
        key=lambda x: x['fecha'],
        reverse=True
    )

    ventas_por_mes = sorted(
        [{'mes': k, **v} for k, v in ventas_mensuales.items()],
        key=lambda x: x['mes'],
        reverse=True
    )

    return render_template(
        'reporte_periodo.html',
        ventas_por_dia=ventas_por_dia,
        ventas_por_mes=ventas_por_mes
    )


@app.route("/reportes/inventario-total")
@login_required
def reporte_inventario_total():
    """Reporte completo del inventario total"""
    productos = cargar_productos()

    if not productos:
        flash("No hay productos en el inventario para mostrar el reporte.", "info")
        return redirect(url_for('lista_productos'))

    total_items = len(productos)
    total_unidades = sum(p.get('stock', 0) for p in productos)
    valor_total_inventario = sum(p.get('stock', 0) * p.get('precio_unitario', 0) for p in productos)

    productos_ordenados = sorted(productos, key=lambda x: x.get('id', 0))

    return render_template(
        "reportes/reporte_inventario_total.html",
        productos=productos_ordenados,
        total_items=total_items,
        total_unidades=total_unidades,
        valor_total=valor_total_inventario
    )


@app.route("/reportes/inventario-total/excel")
@login_required
def exportar_excel_inventario():
    """Exporta el inventario total a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        flash("openpyxl no está disponible. Instala 'openpyxl' para exportar a Excel.", "error")
        return redirect(url_for('reporte_inventario_total'))

    import io

    productos = cargar_productos()
    productos_sorted = sorted(productos, key=lambda x: x.get('id', 0))

    total_items = len(productos_sorted)
    total_unidades = sum(p.get('stock', 0) for p in productos_sorted)
    valor_total = sum(p.get('stock', 0) * p.get('precio_unitario', 0) for p in productos_sorted)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario Total"

    ws.merge_cells('A1:F1')
    cell_title = ws['A1']
    cell_title.value = "Reporte de Inventario Total - Motorrepuestos H&D"
    cell_title.font = Font(size=14, bold=True)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A2:B2')
    ws.merge_cells('C2:D2')
    ws.merge_cells('E2:F2')

    ws['A2'].value = "Total de Productos"
    ws['C2'].value = "Total de Unidades"
    ws['E2'].value = "Valor Total del Inventario"

    ws['A2'].font = Font(bold=True)
    ws['C2'].font = Font(bold=True)
    ws['E2'].font = Font(bold=True)

    ws.merge_cells('A3:B3')
    ws.merge_cells('C3:D3')
    ws.merge_cells('E3:F3')

    ws['A3'].value = total_items
    ws['C3'].value = total_unidades
    ws['E3'].value = valor_total

    ws['A3'].alignment = Alignment(horizontal="center")
    ws['C3'].alignment = Alignment(horizontal="center")
    ws['E3'].alignment = Alignment(horizontal="center")

    box_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for cell_coord in ['A2', 'C2', 'E2', 'A3', 'C3', 'E3']:
        ws[cell_coord].fill = box_fill

    start_row = 5
    encabezados = ["ID", "Producto", "Categoría", "Cantidad", "Precio Unitario", "Valor Total"]

    for col_idx, header in enumerate(encabezados, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = start_row + 1
    for p in productos_sorted:
        cantidad = p.get("stock", 0)
        precio = p.get("precio_unitario", 0)
        valor = round(cantidad * precio, 2)

        ws.cell(row=row, column=1, value=p.get("id"))
        ws.cell(row=row, column=2, value=p.get("nombre"))
        ws.cell(row=row, column=3, value=p.get("categoria"))
        ws.cell(row=row, column=4, value=cantidad)

        cell_precio = ws.cell(row=row, column=5, value=precio)
        cell_valor = ws.cell(row=row, column=6, value=valor)

        cell_precio.number_format = '"$"#,##0.00'
        cell_valor.number_format = '"$"#,##0.00'

        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        row += 1

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    max_col = len(encabezados)
    max_row = ws.max_row

    for r in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=max_col):
        for c in r:
            c.border = thin_border

    for idx, column_cells in enumerate(ws.columns, start=1):
        try:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        except:
            length = 10

        width = min(max(length + 2, 10), 60)
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="inventario_total.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/reportes/inventario-total/pdf")
@login_required
def exportar_pdf_inventario():
    """Exporta el inventario total a PDF"""
    try:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
    except Exception:
        flash("reportlab no está disponible. Instala 'reportlab' para exportar a PDF.", "error")
        return redirect(url_for('reporte_inventario_total'))

    productos = cargar_productos()
    productos_sorted = sorted(productos, key=lambda x: x.get('id', 0))

    total_items = len(productos_sorted)
    total_unidades = sum(p.get('stock', 0) for p in productos_sorted)
    valor_total = sum(p.get('stock', 0) * p.get('precio_unitario', 0) for p in productos_sorted)

    import io
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        name="TitleCenter",
        parent=styles['Heading1'],
        alignment=1,
        fontSize=18,
        spaceAfter=8
    )

    story.append(Paragraph("Reporte de Inventario Total - Motorrepuestos H&D", title_style))
    story.append(Spacer(1, 8))

    totals_data = [
        ["Total de Productos", str(total_items)],
        ["Total de Unidades", str(total_unidades)],
        ["Valor Total del Inventario", f"${valor_total:,.2f}"]
    ]
    totals_table = Table(totals_data, colWidths=[200, 120], hAlign='CENTER')
    totals_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 12))

    data = [["ID", "Producto", "Categoría", "Cantidad", "Precio Unitario", "Valor Total"]]
    for p in productos_sorted:
        cantidad = p.get("stock", 0)
        precio = p.get("precio_unitario", 0)
        valor = round(cantidad * precio, 2)

        data.append([
            p.get("id"),
            p.get("nombre"),
            p.get("categoria"),
            str(cantidad),
            f"${precio:,.2f}",
            f"${valor:,.2f}"
        ])

    col_widths = [40, 240, 120, 60, 90, 100]

    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign='CENTER')

    table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
    ])

    table.setStyle(table_style)
    story.append(table)

    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="inventario_total.pdf",
        mimetype="application/pdf"
    )

@app.route('/reportes/iva')
@login_required
def reporte_iva():
    """Reporte de IVA a pagar al gobierno"""
    anio_filtro = request.args.get('anio', '')
    mes_filtro = request.args.get('mes', '')
    
    query = """
        SELECT 
            YEAR(fecha) as anio,
            MONTH(fecha) as mes,
            COUNT(*) as num_ventas,
            SUM(total) as total_vendido,
            SUM(iva_total) as iva_total
        FROM ventas
        WHERE 1=1
    """
    
    params = []
    filtro_texto = ""
    
    if anio_filtro:
        query += " AND YEAR(fecha) = %s"
        params.append(anio_filtro)
        filtro_texto = f"Año {anio_filtro}"
    
    if mes_filtro:
        query += " AND MONTH(fecha) = %s"
        params.append(mes_filtro)
        meses = {
            '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
            '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
            '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
        }
        if filtro_texto:
            filtro_texto += f" - {meses.get(mes_filtro, 'Mes ' + mes_filtro)}"
        else:
            filtro_texto = meses.get(mes_filtro, 'Mes ' + mes_filtro)
    
    query += " GROUP BY YEAR(fecha), MONTH(fecha) ORDER BY anio DESC, mes DESC"
    
    resultados = ejecutar_query(query, tuple(params) if params else None, fetch_all=True)
    
    iva_por_mes = []
    total_iva = 0
    total_vendido = 0
    total_ventas = 0
    
    meses_nombres = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    if resultados:
        for r in resultados:
            iva_por_mes.append({
                'anio': r['anio'],
                'mes': r['mes'],
                'mes_nombre': meses_nombres.get(r['mes'], str(r['mes'])),
                'num_ventas': r['num_ventas'],
                'total_vendido': r['total_vendido'] or 0,
                'iva_total': r['iva_total'] or 0
            })
            total_iva += r['iva_total'] or 0
            total_vendido += r['total_vendido'] or 0
            total_ventas += r['num_ventas']
    
    query_anios = "SELECT DISTINCT YEAR(fecha) as anio FROM ventas ORDER BY anio DESC"
    anios_result = ejecutar_query(query_anios, fetch_all=True)
    anios = [str(r['anio']) for r in anios_result] if anios_result else []
    
    return render_template(
        'reportes/reporte_iva.html',
        iva_por_mes=iva_por_mes,
        total_iva=total_iva,
        total_vendido=total_vendido,
        total_ventas=total_ventas,
        anios=anios,
        filtro_texto=filtro_texto
    )


@app.route('/reportes/iva/excel')
@login_required
def exportar_iva_excel():
    """Exporta el reporte de IVA a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        flash("openpyxl no está disponible. Instala 'openpyxl' para exportar a Excel.", "error")
        return redirect(url_for('reporte_iva'))

    import io
    
    anio_filtro = request.args.get('anio', '')
    mes_filtro = request.args.get('mes', '')
    
    query = """
        SELECT 
            YEAR(fecha) as anio,
            MONTH(fecha) as mes,
            COUNT(*) as num_ventas,
            SUM(total) as total_vendido,
            SUM(iva_total) as iva_total
        FROM ventas
        WHERE 1=1
    """
    
    params = []
    
    if anio_filtro:
        query += " AND YEAR(fecha) = %s"
        params.append(anio_filtro)
    
    if mes_filtro:
        query += " AND MONTH(fecha) = %s"
        params.append(mes_filtro)
    
    query += " GROUP BY YEAR(fecha), MONTH(fecha) ORDER BY anio DESC, mes DESC"
    
    resultados = ejecutar_query(query, tuple(params) if params else None, fetch_all=True)
    
    meses_nombres = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte IVA"
    
    ws.merge_cells('A1:E1')
    cell_title = ws['A1']
    cell_title.value = "Reporte de IVA a Pagar al Gobierno"
    cell_title.font = Font(size=16, bold=True, color="FFFFFF")
    cell_title.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    total_iva = sum(r['iva_total'] or 0 for r in resultados) if resultados else 0
    total_vendido = sum(r['total_vendido'] or 0 for r in resultados) if resultados else 0
    total_ventas = sum(r['num_ventas'] for r in resultados) if resultados else 0
    
    ws.merge_cells('A3:B3')
    ws.merge_cells('C3:D3')
    ws['A3'] = "Total Ventas Realizadas"
    ws['C3'] = "Total Vendido"
    ws['E3'] = "IVA a Pagar"
    
    ws['A3'].font = Font(bold=True)
    ws['C3'].font = Font(bold=True)
    ws['E3'].font = Font(bold=True)
    
    ws.merge_cells('A4:B4')
    ws.merge_cells('C4:D4')
    ws['A4'] = total_ventas
    ws['C4'] = total_vendido
    ws['E4'] = total_iva
    
    ws['A4'].alignment = Alignment(horizontal="center")
    ws['C4'].alignment = Alignment(horizontal="center")
    ws['E4'].alignment = Alignment(horizontal="center")
    
    ws['C4'].number_format = '"$"#,##0.00'
    ws['E4'].number_format = '"$"#,##0.00'
    ws['E4'].font = Font(bold=True, size=14, color="E67E22")
    
    summary_fill = PatternFill(start_color="FEF5E7", end_color="FEF5E7", fill_type="solid")
    for cell in ['A3', 'C3', 'E3', 'A4', 'C4', 'E4']:
        ws[cell].fill = summary_fill
    
    start_row = 6
    encabezados = ["Año", "Mes", "Ventas Realizadas", "Total Vendido", "IVA Cobrado"]
    
    for col_idx, header in enumerate(encabezados, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row = start_row + 1
    if resultados:
        for r in resultados:
            ws.cell(row=row, column=1, value=r['anio'])
            ws.cell(row=row, column=2, value=meses_nombres.get(r['mes'], str(r['mes'])))
            ws.cell(row=row, column=3, value=r['num_ventas'])
            
            cell_total = ws.cell(row=row, column=4, value=r['total_vendido'] or 0)
            cell_total.number_format = '"$"#,##0.00'
            
            cell_iva = ws.cell(row=row, column=5, value=r['iva_total'] or 0)
            cell_iva.number_format = '"$"#,##0.00'
            cell_iva.font = Font(color="E67E22", bold=True)
            
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            
            row += 1
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    max_row = row - 1
    for r in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=5):
        for cell in r:
            cell.border = thin_border
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name="reporte_iva.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/reportes/rentabilidad')
@login_required
def reporte_rentabilidad():
    """Reporte de productos más rentables"""
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    ordenar = request.args.get('ordenar', 'ganancia')
    
    query = """
        SELECT 
            v.producto_id,
            p.nombre,
            p.categoria,
            COUNT(*) as num_ventas,
            SUM(v.cantidad) as cantidad_vendida,
            SUM(v.total) as total_vendido,
            AVG(v.ganancia_unitaria) as ganancia_unitaria_promedio,
            SUM(v.ganancia_total) as ganancia_total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        WHERE 1=1
    """
    
    params = []
    
    if fecha_desde:
        query += " AND DATE(v.fecha) >= %s"
        params.append(fecha_desde)
    
    if fecha_hasta:
        query += " AND DATE(v.fecha) <= %s"
        params.append(fecha_hasta)
    
    query += " GROUP BY v.producto_id, p.nombre, p.categoria"
    
    if ordenar == 'cantidad':
        query += " ORDER BY cantidad_vendida DESC"
    else:
        query += " ORDER BY ganancia_total DESC"
    
    resultados = ejecutar_query(query, tuple(params) if params else None, fetch_all=True)
    
    productos_rentables = []
    ganancia_total = 0
    unidades_vendidas = 0
    
    if resultados:
        for r in resultados:
            ganancia_tot = r['ganancia_total'] or 0
            total_vend = r['total_vendido'] or 0
            
            if total_vend > 0:
                margen = (ganancia_tot / total_vend) * 100
            else:
                margen = 0
            
            productos_rentables.append({
                'producto_id': r['producto_id'],
                'nombre': r['nombre'],
                'categoria': r['categoria'],
                'num_ventas': r['num_ventas'],
                'cantidad_vendida': r['cantidad_vendida'],
                'total_vendido': total_vend,
                'ganancia_unitaria_promedio': r['ganancia_unitaria_promedio'] or 0,
                'ganancia_total': ganancia_tot,
                'margen_porcentaje': margen
            })
            
            ganancia_total += ganancia_tot
            unidades_vendidas += r['cantidad_vendida']
    
    query_cat = """
        SELECT 
            p.categoria,
            COUNT(DISTINCT v.producto_id) as num_productos,
            SUM(v.cantidad) as unidades_vendidas,
            SUM(v.ganancia_total) as ganancia_total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        WHERE 1=1
    """
    
    params_cat = []
    
    if fecha_desde:
        query_cat += " AND DATE(v.fecha) >= %s"
        params_cat.append(fecha_desde)
    
    if fecha_hasta:
        query_cat += " AND DATE(v.fecha) <= %s"
        params_cat.append(fecha_hasta)
    
    query_cat += " GROUP BY p.categoria ORDER BY ganancia_total DESC"
    
    resultados_cat = ejecutar_query(query_cat, tuple(params_cat) if params_cat else None, fetch_all=True)
    
    rentabilidad_categorias = []
    if resultados_cat:
        for r in resultados_cat:
            ganancia_cat = r['ganancia_total'] or 0
            num_prod = r['num_productos']
            
            rentabilidad_categorias.append({
                'categoria': r['categoria'],
                'num_productos': num_prod,
                'unidades_vendidas': r['unidades_vendidas'],
                'ganancia_total': ganancia_cat,
                'ganancia_promedio': ganancia_cat / num_prod if num_prod > 0 else 0
            })
    
    return render_template(
        'reportes/reporte_rentabilidad.html',
        productos_rentables=productos_rentables,
        rentabilidad_categorias=rentabilidad_categorias,
        ganancia_total=ganancia_total,
        total_productos=len(productos_rentables),
        unidades_vendidas=unidades_vendidas
    )


@app.route('/reportes/rentabilidad/excel')
@login_required
def exportar_rentabilidad_excel():
    """Exporta el reporte de rentabilidad a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        flash("openpyxl no está disponible.", "error")
        return redirect(url_for('reporte_rentabilidad'))

    import io
    
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    ordenar = request.args.get('ordenar', 'ganancia')
    
    query = """
        SELECT 
            v.producto_id,
            p.nombre,
            p.categoria,
            COUNT(*) as num_ventas,
            SUM(v.cantidad) as cantidad_vendida,
            SUM(v.total) as total_vendido,
            AVG(v.ganancia_unitaria) as ganancia_unitaria_promedio,
            SUM(v.ganancia_total) as ganancia_total
        FROM ventas v
        JOIN productos p ON v.producto_id = p.id
        WHERE 1=1
    """
    
    params = []
    
    if fecha_desde:
        query += " AND DATE(v.fecha) >= %s"
        params.append(fecha_desde)
    
    if fecha_hasta:
        query += " AND DATE(v.fecha) <= %s"
        params.append(fecha_hasta)
    
    query += " GROUP BY v.producto_id, p.nombre, p.categoria"
    
    if ordenar == 'cantidad':
        query += " ORDER BY cantidad_vendida DESC"
    else:
        query += " ORDER BY ganancia_total DESC"
    
    resultados = ejecutar_query(query, tuple(params) if params else None, fetch_all=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rentabilidad"
    
    ws.merge_cells('A1:H1')
    cell_title = ws['A1']
    cell_title.value = "Reporte de Productos Más Rentables"
    cell_title.font = Font(size=16, bold=True, color="FFFFFF")
    cell_title.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    ganancia_total = sum(r['ganancia_total'] or 0 for r in resultados) if resultados else 0
    unidades_vendidas = sum(r['cantidad_vendida'] for r in resultados) if resultados else 0
    
    ws.merge_cells('A3:B3')
    ws.merge_cells('C3:D3')
    ws.merge_cells('E3:F3')
    ws['A3'] = "Ganancia Total"
    ws['C3'] = "Productos Vendidos"
    ws['E3'] = "Unidades Vendidas"
    
    ws['A3'].font = Font(bold=True)
    ws['C3'].font = Font(bold=True)
    ws['E3'].font = Font(bold=True)
    
    ws.merge_cells('A4:B4')
    ws.merge_cells('C4:D4')
    ws.merge_cells('E4:F4')
    ws['A4'] = ganancia_total
    ws['C4'] = len(resultados)
    ws['E4'] = unidades_vendidas
    
    ws['A4'].alignment = Alignment(horizontal="center")
    ws['C4'].alignment = Alignment(horizontal="center")
    ws['E4'].alignment = Alignment(horizontal="center")
    
    ws['A4'].number_format = '"$"#,##0.00'
    ws['A4'].font = Font(bold=True, size=14, color="27AE60")
    
    summary_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
    for cell in ['A3', 'C3', 'E3', 'A4', 'C4', 'E4']:
        ws[cell].fill = summary_fill
    
    start_row = 6
    encabezados = ["Posición", "Producto", "Categoría", "Unidades", "Total Vendido", "Gan. Unitaria", "Gan. Total", "% Margen"]
    
    for col_idx, header in enumerate(encabezados, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row = start_row + 1
    if resultados:
        for idx, r in enumerate(resultados, start=1):
            ganancia_tot = r['ganancia_total'] or 0
            total_vend = r['total_vendido'] or 0
            margen = (ganancia_tot / total_vend * 100) if total_vend > 0 else 0
            
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=r['nombre'])
            ws.cell(row=row, column=3, value=r['categoria'])
            ws.cell(row=row, column=4, value=r['cantidad_vendida'])
            
            cell_total = ws.cell(row=row, column=5, value=total_vend)
            cell_total.number_format = '"$"#,##0.00'
            
            cell_gan_unit = ws.cell(row=row, column=6, value=r['ganancia_unitaria_promedio'] or 0)
            cell_gan_unit.number_format = '"$"#,##0.00'
            
            cell_gan_total = ws.cell(row=row, column=7, value=ganancia_tot)
            cell_gan_total.number_format = '"$"#,##0.00'
            cell_gan_total.font = Font(bold=True, color="27AE60")
            
            cell_margen = ws.cell(row=row, column=8, value=margen)
            cell_margen.number_format = '0.0"%"'
            
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            
            if idx <= 3:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFF9E6", end_color="FFF9E6", fill_type="solid"
                    )
            
            row += 1
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    max_row = row - 1
    for r in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=8):
        for cell in r:
            cell.border = thin_border
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name="reporte_rentabilidad.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('500.html'), 500


@app.route("/cargar_excel")
def cargar_excel():
    return render_template("cargar_excel.html")


import openpyxl

@app.route("/importar_excel", methods=["POST"])
def importar_excel():
    if "excel_file" not in request.files:
        flash("No se subió ningún archivo", "danger")
        return redirect(url_for("cargar_excel"))

    file = request.files["excel_file"]

    # ========== CREAR RUTA uploads ==========
    import os
    UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads")

    # si no existe la carpeta, crearla
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)

    # ruta final del archivo
    filepath = os.path.join(UPLOAD_FOLDER, "productos.xlsx")

    # guardar archivo
    file.save(filepath)

    # leer excel
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    # recorrer filas
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nombre, categoria, marca, stock, precio_unitario, sku, descripcion = row

        if not nombre:
            continue

        query = """
            INSERT INTO productos (nombre, categoria, marca, stock, precio_unitario, sku, descripcion)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """

        params = (nombre, categoria, marca, stock, precio_unitario, sku, descripcion)

        ejecutar_query(query, params, commit=True)

    flash("Productos importados correctamente desde Excel", "success")
    return redirect(url_for("cargar_excel"))

# ---------------------------------------------------------------------------------
# RUN
# ---------------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True)